"""
Rellena la plantilla Excel 'Cartón médico.xlsx' con los datos parseados.
"""
import io
import os

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment
from openpyxl.worksheet.page import PageMargins

from config import TEMPLATE_PATH
from app.excel.formatter import escribir_en_una_linea, normalizar_interrogatorio
from app.utils.helpers import separar_apellido_nombre, formatear_dosis


def completar_carton_medico(datos):
    """
    Abre 'Cartón médico.xlsx', rellena campos en la hoja 'Frente'
    y devuelve el contenido del xlsx como BytesIO.
    """
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(
            f"No se encontró la plantilla Cartón médico.xlsx en {TEMPLATE_PATH}"
        )

    wb = load_workbook(TEMPLATE_PATH)

    # Márgenes en CERO para que el PDF salga sin bordes
    margins = PageMargins(left=0, right=0, top=0, bottom=0, header=0, footer=0)
    for hoja in wb.worksheets:
        hoja.page_margins = margins

    ws_f = wb["Frente"]

    # --- Datos del paciente ---
    nombre_completo = datos.get("nombre_completo")
    edad = datos.get("edad")

    apellido, nombre = separar_apellido_nombre(nombre_completo)

    if apellido:
        escribir_en_una_linea(ws_f, "C9", apellido, horizontal="center")

    if nombre:
        escribir_en_una_linea(ws_f, "C10", nombre, horizontal="center")

    if edad is not None:
        ws_f["I9"] = edad

    # Peso corporal
    peso = datos.get("peso")
    if peso is not None:
        ws_f["I10"] = f"{peso} kg"

    # ID del paciente
    id_paciente = datos.get("id")
    if id_paciente:
        ws_f["C11"] = id_paciente

    # --- Diagnóstico ---
    if datos.get("diagnostico"):
        ws_f["C14"] = datos["diagnostico"]

    # --- Estadificación ---
    if datos.get("estad_t"):
        ws_f["C15"] = f"T: {datos['estad_t']}"

    if datos.get("estad_n"):
        ws_f["D15"] = f"N: {datos['estad_n']}"

    if datos.get("estad_m"):
        ws_f["E15"] = f"M: {datos['estad_m']}"

    if datos.get("estad_estadio"):
        ws_f["G15"] = f"ESTADIO: {datos['estad_estadio']}"

    # --- Histología ---
    if datos.get("histologia"):
        ws_f["C16"] = datos["histologia"]

    # --- Interrogatorio (C17) ---
    if datos.get("interrogatorio"):
        texto_inter = normalizar_interrogatorio(datos["interrogatorio"])
        cell = ws_f["C17"]

        if isinstance(cell, MergedCell):
            for merged_range in ws_f.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left = ws_f.cell(
                        row=merged_range.min_row, column=merged_range.min_col
                    )
                    top_left.value = texto_inter
                    top_left.alignment = Alignment(
                        wrap_text=True, vertical="top", horizontal="left"
                    )
                    break
        else:
            cell.value = texto_inter
            cell.alignment = Alignment(
                wrap_text=True, vertical="top", horizontal="left"
            )

    # --- Prescripción braquiterapia ---
    ws_d = wb["Dorso"]  # noqa: F841 (referencia para uso futuro)

    dosis_total = datos.get("braqui_dosis_total")
    n_fx = datos.get("braqui_n_fracciones")
    dosis_fx = datos.get("braqui_dosis_por_fraccion")

    if dosis_total is not None:
        ws_f["C36"] = formatear_dosis(dosis_total)

    if n_fx is not None:
        ws_f["C37"] = int(n_fx) if float(n_fx).is_integer() else n_fx

    if dosis_fx is not None:
        ws_f["C38"] = formatear_dosis(dosis_fx)

    # Guardar a un buffer en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
